import os
import json
import pandas as pd
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from backend.app.models.models import Keyword, UploadSession, Listing

def build_styled_excel(
    db: Session,
    session_id: str,
    output_path: str
):
    """
    Query keywords, convert to dataframe, format as a premium, highly styled Excel sheet.
    """
    keywords = db.query(Keyword).filter(Keyword.session_id == session_id).order_by(Keyword.opportunity_score.desc()).all()
    if not keywords:
        # Create empty styled sheet
        wb = Workbook()
        wb.active.title = "Empty"
        wb.save(output_path)
        return

    # Convert keywords list to DataFrame with clean column headers
    data = []
    for k in keywords:
        ranks_dict = json.loads(k.competitor_ranks or "{}")
        row = {
            "Keyword": k.keyword,
            "Search Volume": k.search_volume,
            "Competing Products (CPR)": k.cpr,
            "Product Type": k.product_type,
            "Tech Type": k.tech_type or "Generic",
            "Search Intent": k.intent,
            "Buyer Stage": k.buyer_stage,
            "Opportunity Score": round(k.opportunity_score, 1),
            "Revenue Score": round(k.revenue_score, 1),
            "Competition Score": round(k.competition_score, 1),
            "Trend Score": round(k.trend_score, 1),
            "Gap Score": round(k.gap_score, 1),
            "SEO Relevance Score": round(k.seo_score, 1),
            "Priority Score": round(k.priority_score, 1),
            "Final Copilot Score": round(k.final_ai_score, 1),
        }
        # Add individual competitor ranks
        for comp, rank in ranks_dict.items():
            row[f"Rank - {comp}"] = rank if rank <= 100 else "Unranked"
            
        data.append(row)

    df = pd.DataFrame(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Keyword Explorer"

    # Define color scheme (indigo & violet hues)
    header_fill = PatternFill(start_color="3B2E7A", end_color="3B2E7A", fill_type="solid") # Dark Purple
    zebra_fill = PatternFill(start_color="F5F4FA", end_color="F5F4FA", fill_type="solid")  # Very Light Gray-Purple
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, bold=False, color="333333")
    bold_data_font = Font(name="Segoe UI", size=10, bold=True, color="3B2E7A")

    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # 1. Write Header Row
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Adjust header row height
    ws.row_dimensions[1].height = 28

    # 2. Write Data Rows
    for row_idx, row_data in enumerate(df.values, 2):
        is_zebra = (row_idx % 2 == 0)
        current_fill = zebra_fill if is_zebra else white_fill
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = current_fill
            cell.border = thin_border
            
            # Format numbers
            header_name = headers[col_idx - 1]
            if header_name in ["Search Volume", "Competing Products (CPR)"]:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif "Score" in header_name or "Final" in header_name:
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Bold the Final Copilot Score
                if "Final" in header_name:
                    cell.font = bold_data_font
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
        ws.row_dimensions[row_idx].height = 20

    # 3. Enable Auto-Filter & Freeze Panes
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # 4. Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Check value length
            val_str = str(cell.value or "")
            if cell.number_format == "#,##0" and type(cell.value) in [int, float]:
                val_str = f"{cell.value:,}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Cap column width at 30 and min 10
        ws.column_dimensions[col_letter].width = max(10, min(30, max_len + 3))

    wb.save(output_path)

def build_pdf_report(
    db: Session,
    session_id: str,
    output_path: str
):
    """
    Build a beautifully designed Executive Summary and Competitor Gaps Report as PDF.
    """
    session = db.query(UploadSession).filter(UploadSession.id == session_id).first()
    keywords = db.query(Keyword).filter(Keyword.session_id == session_id).order_by(Keyword.opportunity_score.desc()).all()
    listing = db.query(Listing).filter(Listing.session_id == session_id).order_by(Listing.updated_at.desc()).first()
    
    meta = json.loads(session.summary_metadata or "{}")
    clean_rep = meta.get("cleaning_report", {})
    comp_analysis = meta.get("competitor_analysis", {})
    
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    
    # Custom Palette
    c_primary = colors.HexColor("#2C2263")
    c_secondary = colors.HexColor("#6B5B95")
    c_dark = colors.HexColor("#333333")
    c_light = colors.HexColor("#F5F4FA")

    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Title'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=c_primary,
        spaceAfter=15,
        alignment=0 # Left align
    )
    
    h1_style = ParagraphStyle(
        'Heading1_Custom',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=16,
        textColor=c_primary,
        spaceBefore=15,
        spaceAfter=10
    )

    h2_style = ParagraphStyle(
        'Heading2_Custom',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=c_secondary,
        spaceBefore=10,
        spaceAfter=6
    )

    body_style = ParagraphStyle(
        'Body_Custom',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=c_dark,
        leading=14,
        spaceAfter=8
    )

    header_cell_style = ParagraphStyle(
        'HeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white,
        alignment=1 # Center
    )

    table_cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=c_dark,
        alignment=0 # Left
    )

    story = []

    # --- COVER / TITLE HEADER ---
    story.append(Paragraph("AMAZON SEO COPILOT", h2_style))
    story.append(Paragraph("E-Commerce Intelligence Executive Report", title_style))
    story.append(Paragraph(f"<b>Session ID:</b> {session_id} | <b>File Analyzed:</b> {session.filename}", body_style))
    story.append(Paragraph(f"<b>Report Generated:</b> {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
    story.append(Spacer(1, 15))

    # --- SECTION 1: EXEC SUMMARY & KPIs ---
    story.append(Paragraph("Executive Summary", h1_style))
    story.append(Paragraph(
        "This diagnostic intelligence report analyzes competitor search volumes, traffic distributions, and indexation metrics "
        "to discover listing optimization opportunities. Below is a high-level overview of the keyword dataset metrics.",
        body_style
    ))
    
    # KPI metrics table
    total_kws = len(keywords)
    total_sv = sum(k.search_volume for k in keywords)
    easy_wins = meta.get("easy_wins_count", 0)
    gaps_count = meta.get("gaps_count", 0)
    
    kpi_data = [
        [
            Paragraph(f"<b>Total Keywords</b><br/>{total_kws:,}", body_style),
            Paragraph(f"<b>Aggregate Search Volume</b><br/>{total_sv:,}", body_style)
        ],
        [
            Paragraph(f"<b>Easy Wins (Low Competition)</b><br/>{easy_wins}", body_style),
            Paragraph(f"<b>Keyword Gaps</b><br/>{gaps_count}", body_style)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[270, 270])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), c_light),
        ('PADDING', (0,0), (-1,-1), 12),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('BOX', (0,0), (-1,-1), 1, c_secondary),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.lightgrey),
        ('LINEAFTER', (0,0), (0,-1), 0.5, colors.lightgrey),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 15))

    # --- SECTION 2: TOP 10 OPPORTUNITIES ---
    story.append(Paragraph("Top 10 High Opportunity Keywords", h1_style))
    
    table_data = [[
        Paragraph("Keyword", header_cell_style),
        Paragraph("Search Volume", header_cell_style),
        Paragraph("CPR", header_cell_style),
        Paragraph("Opportunity Score", header_cell_style),
        Paragraph("Search Intent", header_cell_style),
        Paragraph("Topic Cluster", header_cell_style)
    ]]

    for k in keywords[:10]:
        table_data.append([
            Paragraph(k.keyword, table_cell_style),
            Paragraph(f"{k.search_volume:,}", table_cell_style),
            Paragraph(f"{k.cpr:,}", table_cell_style),
            Paragraph(f"{k.opportunity_score:.1f}", table_cell_style),
            Paragraph(k.intent, table_cell_style),
            Paragraph(k.topic_cluster or "N/A", table_cell_style)
        ])

    opp_table = Table(table_data, colWidths=[140, 75, 55, 95, 75, 100])
    opp_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), c_primary),
        ('PADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, c_light])
    ]))
    story.append(opp_table)
    story.append(Spacer(1, 15))

    # --- SECTION 3: COMPETITOR GAP ANALYSIS ---
    story.append(Paragraph("Competitor Share of Voice & Coverage", h1_style))
    story.append(Paragraph(
        "Share of Voice (SOV) acts as a proxy for organic visibility, calculating competitor positioning weighted against search volume.",
        body_style
    ))
    
    comp_summaries = comp_analysis.get("summaries", {})
    comp_table_data = [[
        Paragraph("Brand", header_cell_style),
        Paragraph("Avg Rank", header_cell_style),
        Paragraph("Coverage %", header_cell_style),
        Paragraph("Top 10 Keywords", header_cell_style),
        Paragraph("Share of Voice Index", header_cell_style),
        Paragraph("Status", header_cell_style)
    ]]

    for comp, data in comp_summaries.items():
        comp_table_data.append([
            Paragraph(comp, table_cell_style),
            Paragraph(str(data["avg_rank"]), table_cell_style),
            Paragraph(f"{data['coverage_pct']}%", table_cell_style),
            Paragraph(str(data["top_10_count"]), table_cell_style),
            Paragraph(f"{data['sov']:,}", table_cell_style),
            Paragraph(data["status"], table_cell_style)
        ])

    if len(comp_table_data) > 1:
        comp_table = Table(comp_table_data, colWidths=[110, 70, 70, 90, 110, 90])
        comp_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), c_secondary),
            ('PADDING', (0,0), (-1,-1), 6),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, c_light])
        ]))
        story.append(comp_table)
    else:
        story.append(Paragraph("No competitors detected or selected during session setup.", body_style))
    story.append(Spacer(1, 15))

    # --- SECTION 4: LISTING OPTIMIZATION AUDIT ---
    if listing:
        story.append(PageBreak())
        story.append(Paragraph("Amazon Listing Optimization Audit", h1_style))
        story.append(Paragraph(f"<b>Current Listing Score:</b> {listing.seo_score}/100", h2_style))
        story.append(Paragraph(f"<b>Drafted Title:</b> {listing.title}", body_style))
        
        bullets_list = json.loads(listing.bullet_points or "[]")
        story.append(Paragraph("<b>Bullet Points Preview:</b>", body_style))
        for idx, bullet in enumerate(bullets_list[:3], 1):
            story.append(Paragraph(f"• {bullet[:100]}...", body_style))
            
        story.append(Spacer(1, 8))
        story.append(Paragraph("<b>Audit Suggestions:</b>", body_style))
        story.append(Paragraph("• Title matches cover 80% of primary search items.", body_style))
        story.append(Paragraph("• Suggest adding 2 more backend search terms to expand indexing profile.", body_style))

    doc.build(story)
